"""PDF and Excel export for the Data Explorer table.
PDF: matplotlib PdfPages with paginated A4 tables (no fpdf2 dependency).
Excel: pandas + openpyxl (standard, robust)."""
import io
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages


def export_excel(df_display: pd.DataFrame, year: int, lang: str) -> bytes:
    """Export to a clean .xlsx file with one sheet."""
    sheet = f"Year_{year}"
    buf = io.BytesIO()
    try:
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_display.to_excel(writer, sheet_name=sheet, index=False)
            from openpyxl.styles import Font, PatternFill
            ws = writer.sheets[sheet]
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_fill = PatternFill(
                start_color="0B2540",
                end_color="0B2540",
                fill_type="solid")
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
            for column in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 35)
    except Exception:
        df_display.to_excel(buf, sheet_name=sheet, index=False)
    buf.seek(0)
    return buf.getvalue()


def export_pdf(df_display: pd.DataFrame, year: int, lang: str) -> bytes:
    """Export to a paginated A4 PDF (cover page + data tables)."""
    buf = io.BytesIO()
    title = ("World Economic Data" if lang != "fr" else "Donnees Economiques Mondiales")
    subtitle = (f"{len(df_display)} countries - Year {year}" if lang != "fr"
                else f"{len(df_display)} pays - Annee {year}")

    rows_per_page = 28
    cols_per_page = 7
    cols = df_display.columns.tolist()
    str_df = df_display.astype(str).replace({"nan": "-", "None": "-", "": "-"})

    with PdfPages(buf) as pdf:
        fig = plt.figure(figsize=(8.27, 11.69))
        ax = fig.add_subplot(111)
        ax.axis("off")
        ax.text(0.5, 0.62, title, ha="center", va="center", fontsize=22, fontweight="bold", color="#0B2540")
        ax.text(0.5, 0.55, subtitle, ha="center", va="center", fontsize=13, color="#475569")
        ax.text(0.5, 0.48, "Global Economic Intelligence Dashboard", ha="center", va="center", fontsize=10, color="#94a3b8", style="italic")
        pdf.savefig(fig)
        plt.close(fig)

        col_pages = [cols[i:i + cols_per_page] for i in range(0, len(cols), cols_per_page)]
        for cp_idx, col_page in enumerate(col_pages):
            sub = str_df[col_page]
            for start in range(0, len(sub), rows_per_page):
                chunk = sub.iloc[start:start + rows_per_page]
                fig, ax = plt.subplots(figsize=(8.27, 11.69))
                ax.axis("off")
                page_num = cp_idx * ((len(sub) + rows_per_page - 1) // rows_per_page) + (start // rows_per_page) + 1
                ax.set_title(f"{title} - {year} - Page {page_num}", fontsize=10, color="#0B2540", pad=12)
                table = ax.table(
                    cellText=chunk.values,
                    colLabels=[str(c)[:28] for c in chunk.columns],
                    cellLoc="left",
                    loc="center",
                )
                table.auto_set_font_size(False)
                table.set_fontsize(7)
                table.scale(1, 1.25)
                for j in range(len(chunk.columns)):
                    cell = table[(0, j)]
                    cell.set_facecolor("#0B2540")
                    cell.set_text_props(color="white", fontweight="bold", fontsize=7)
                for i in range(1, len(chunk) + 1):
                    for j in range(len(chunk.columns)):
                        cell = table[(i, j)]
                        if i % 2 == 0:
                            cell.set_facecolor("#f4f9fe")
                        else:
                            cell.set_facecolor("#ffffff")
                pdf.savefig(fig)
                plt.close(fig)
    return buf.getvalue()
